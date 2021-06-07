# -*- coding: utf-8 -*-
"""
Created on Tue Jun  2 14:57:55 2020

@author: harshvardhans

Copy Pastes the items from test.xlsx to GP PO entry Window

"""

import os
import openpyxl
import pyperclip
import pyautogui
import keyboard
import time
import pygetwindow as gw
from win32 import win32gui #brings window to front
import re
import gp_price_format

gp_price_format.num_GP()

user = os.getlogin()

class WindowMgr:
    """Encapsulates some calls to the winapi for window management"""

    def __init__ (self):
        """Constructor"""
        self._handle = None

    def _window_enum_callback(self, hwnd, wildcard):
        """Pass to win32gui.EnumWindows() to check all the opened windows"""
        if re.match(wildcard, str(win32gui.GetWindowText(hwnd))) is not None:
            self._handle = hwnd

    def find_window_wildcard(self, wildcard):
        """find a window whose title matches the wildcard regex"""
        self._handle = None
        win32gui.EnumWindows(self._window_enum_callback, wildcard)

    def set_foreground(self):
        """put the window in the foreground"""
        win32gui.SetForegroundWindow(self._handle)


w = WindowMgr()

try:
    w.find_window_wildcard(".*Purchasing Item Detail Entry.*")
    w.set_foreground()
    time.sleep(1)
    
except:
    while True:
        if ('Purchasing Item Detail Entry - \\\\Remote') in gw.getAllTitles():
            gpWindow = gw.getWindowsWithTitle('Purchasing Item Detail Entry - \\\\Remote')[0]
            gpWindow.restore()
            gpWindow.activate()
            break
        else:
            pyautogui.alert('Open Purchasing Item Detail Entry Window')
            time.sleep (1)

time.sleep(0.2)

wb=openpyxl.load_workbook('C:\\Users\\'+user+'\\Desktop\\python\\test.xlsx', data_only=True)

#for x in range (1,len(wb.sheetnames)):
#    sheet=wb[wb.sheetnames[x]]
sheet = wb.active
max_row=sheet.max_row+1
max_column=sheet.max_column+1

for i in range(3,max_row):
    for j in range(1,max_column):
        if gw.getAllTitles().count('Microsoft Dynamics GP - \\\\Remote') > 1:
            pyautogui.typewrite('password')
            keyboard.press_and_release('\n')
            continue
        cell_obj=sheet.cell(row=i,column=j)
        pyperclip.copy(str(cell_obj.value))
        pyautogui.typewrite(pyperclip.paste())
        keyboard.press_and_release('\t')
        time.sleep(0.5)
    keyboard.press_and_release('\n')
    time.sleep (0.5)

gp_price_format.num_restore()
