import openpyxl
file = 'C:\\Users\\harshvardhans\\Desktop\\python\\test.xlsx'

def num_GP():
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for i in range(3, sheet.max_row+1):
        for j in range(5,7):
            a = sheet.cell(row=i, column=j).value
            a = format(a,'.2f').replace('.','')
            sheet.cell(row=i, column=j).value = a
    wb.save(file)
    

def num_restore():
    ''' Restores the number to correct format '''
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    for i in range(3, sheet.max_row+1):
        for j in range(5,7):
            a = sheet.cell(row=i, column=j).value
            a = int(a)/100
            sheet.cell(row=i, column=j).value = a
    wb.save(file)